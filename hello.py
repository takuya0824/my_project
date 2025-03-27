import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

# サンプルデータの作成
data = {
    'employee_id': range(1001, 1011),  # 従業員ID
    'name': [
        '山田太郎', '鈴木花子', '佐藤一郎', '田中美咲', 
        '中村健一', '小林裕子', '加藤誠', '渡辺春香', 
        '伊藤光男', '高橋恵子'
    ],
    'department': [
        '営業部', '人事部', '技術部', '営業部',
        '技術部', '財務部', '営業部', '人事部',
        '技術部', '財務部'
    ],
    'position': [
        '課長', '主任', '部長', '社員',
        '主任', '課長', '社員', '社員',
        '課長', '主任'
    ],
    'salary': [
        350000, 280000, 450000, 250000,
        300000, 380000, 250000, 250000,
        400000, 290000
    ],
    'hire_date': [
        '2015-04-01', '2018-04-01', '2010-04-01', '2020-04-01',
        '2017-04-01', '2016-04-01', '2021-04-01', '2019-04-01',
        '2012-04-01', '2018-04-01'
    ],
    'email': [
        'yamada.t@example.com', 'suzuki.h@example.com', 'sato.i@example.com',
        'tanaka.m@example.com', 'nakamura.k@example.com', 'kobayashi.y@example.com',
        'kato.m@example.com', 'watanabe.h@example.com', 'ito.m@example.com',
        'takahashi.k@example.com'
    ]
}

# DataFrameの作成
df = pd.DataFrame(data)

# 給与で降順ソート
df = df.sort_values(by='salary', ascending=False)

# Excelに保存 (pandas → openpyxl)
file_path = 'employee_data.xlsx'
df.to_excel(file_path, index=False)

# ---- openpyxl でフォーマット・ソート機能を追加 ----

from openpyxl import load_workbook

# Excelを読み込み
wb = load_workbook(file_path)
ws = wb.active

# ✅ オートフィルタを設定 (ソート機能)
ws.auto_filter.ref = ws.dimensions

# ✅ ユーザー定義のフォーマット設定
# 通貨フォーマット（¥マーク付き）
currency_format = '¥#,##0'

# 日付フォーマット（yyyy-mm-dd）
date_format = 'yyyy-mm-dd'

# 給与列のフォーマット (B2 から B10)
for cell in ws['E'][1:]:  # E列 = 給与
    cell.number_format = currency_format

# 入社日列のフォーマット (F2 から F10)
for cell in ws['F'][1:]:  # F列 = 入社日
    cell.number_format = date_format

# 保存
wb.save(file_path)

# ✅ SELECT文の生成
select_sql = """
SELECT 
    employee_id AS 従業員ID,
    name AS 氏名,
    department AS 部署,
    position AS 役職,
    salary AS 給与,
    hire_date AS 入社日,
    email AS メールアドレス
FROM 
    employees
WHERE 
    department = '営業部'
ORDER BY 
    salary DESC;
"""

# ✅ SQLファイルとして保存
with open('select_query.sql', 'w', encoding='utf-8') as f:
    f.write(select_sql)

print("ExcelファイルとSQLファイルを生成しました。")

#featureブランチでのコメント
